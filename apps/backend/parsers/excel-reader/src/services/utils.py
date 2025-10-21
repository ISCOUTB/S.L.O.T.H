import asyncio
import csv
import time
from functools import wraps
from io import BytesIO
from typing import Any, Callable, Dict, List, TypeVar

import openpyxl

from src.services import dtypes
from src.utils.logger import logger

F = TypeVar("F", bound=Callable[..., Any])


def monitor_performance(operation_name: str):
    def decorator(func: F) -> F:
        if asyncio.iscoroutinefunction(func):

            @wraps(func)
            async def async_wrapper(*args, **kwargs):
                start_time = time.perf_counter()

                try:
                    result = await func(*args, **kwargs)
                    status = "success"
                    return result
                except Exception as e:
                    status = "error"
                    raise e
                finally:
                    end_time = time.perf_counter()
                    duration = end_time - start_time

                    logger.info(
                        f"{operation_name} - Duration: {duration:.4f}s - Status: {status}"
                    )

            return async_wrapper
        else:

            @wraps(func)
            def sync_wrapper(*args, **kwargs):
                start_time = time.perf_counter()

                try:
                    result = func(*args, **kwargs)
                    status = "success"
                    return result
                except Exception as e:
                    status = "error"
                    raise e
                finally:
                    end_time = time.perf_counter()
                    duration = end_time - start_time

                    logger.info(
                        f"{operation_name} - Duration: {duration:.4f}s - Status: {status}"
                    )

            return sync_wrapper

    return decorator


@monitor_performance("open_file_from_bytes")
def open_file_from_bytes(file_bytes: bytes, **kwargs: Any) -> openpyxl.Workbook:
    """
    Open an Excel file from bytes.

    Args:
        file_bytes (bytes): The bytes of the Excel file.
        **kwargs: Additional keyword arguments to pass to openpyxl.load_workbook.

    Returns:
        openpyxl.Workbook: The loaded workbook object.
    """
    excel_file = BytesIO(file_bytes)
    return openpyxl.load_workbook(excel_file, data_only=False, **kwargs)


def extract_formulas(
    workbook: openpyxl.Workbook, limit: int = 50
) -> Dict[str, Dict[str, List[dtypes.CellData]]]:
    """
    Extract formulas and cell data from an Excel workbook.

    Args:
        workbook (openpyxl.Workbook): The workbook object containing the Excel data.
        limit (int): The maximum number of cells to extract per column.

    Returns:
        List[dtypes.CellData]: A list of dictionaries containing cell data, including sheet name,
        cell coordinate, value, data type, and whether the cell contains a formula.
    """
    sheets: Dict[str, Dict[str, List[dtypes.CellData]]] = {}

    for sheet in workbook.worksheets:
        sheets[sheet.title] = {}
        max_rows = (
            sheet.max_column if limit <= 0 else min(limit, sheet.max_column)
        )
        for column in sheet.columns:
            if column[0].value is None:
                continue

            column_letter = column[0].column_letter
            result: List[dtypes.CellData] = []
            for cell, _ in zip(column, range(max_rows)):
                cell_data: dtypes.CellData = {
                    "cell": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    # isinstance(cell.value, str) and cell.value.startswith("=")
                    "is_formula": cell.data_type == "f",
                }
                result.append(cell_data)
            sheets[sheet.title][column_letter] = result

    return sheets


def convert_csv_to_excel(file_bytes: bytes) -> openpyxl.Workbook:
    """
    Convert CSV file bytes to Excel workbook.

    Args:
        file_bytes (bytes): The bytes of the CSV file.

    Returns:
        openpyxl.Workbook: The Excel workbook object.
    """
    csv_text = file_bytes.decode("utf-8")
    csv_reader = csv.reader(csv_text.splitlines())
    csv_data = list(csv_reader)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    for row_idx, row_data in enumerate(csv_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    return workbook
